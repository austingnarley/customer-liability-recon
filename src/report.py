from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.ledger import load_liabilities
from src.prices import load_prices
from src.reconcile import ReconciliationRow

DEFAULT_LEDGER_PATH = Path("data/customer_ledger.csv")
DEFAULT_PRICES_PATH = Path("data/price_snapshots.csv")


def render_html(
    rows: list[ReconciliationRow],
    as_of: date,
    out_path: Path,
    *,
    ledger_path: Path = DEFAULT_LEDGER_PATH,
    prices_path: Path = DEFAULT_PRICES_PATH,
) -> None:
    """Render the single-file HTML report."""
    env = _template_env()
    template = env.get_template("report.html.j2")
    html = template.render(
        rows=rows,
        as_of=as_of.isoformat(),
        generated_at=datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S UTC"),
        top_customers=_top_customers(ledger_path, prices_path),
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(html, encoding="utf-8")


def render_xlsx(
    rows: list[ReconciliationRow],
    as_of: date,
    out_path: Path,
    *,
    ledger_path: Path = DEFAULT_LEDGER_PATH,
    prices_path: Path = DEFAULT_PRICES_PATH,
) -> None:
    """Render the Excel workbook with summary, liabilities, reserves, and methodology sheets."""
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _write_summary(summary, rows, as_of)
    _write_liabilities(workbook.create_sheet("Liabilities"), ledger_path)
    _write_reserves(workbook.create_sheet("Reserves"), rows)
    _write_methodology(workbook.create_sheet("Methodology"))
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        _autosize_columns(sheet)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)


def _template_env() -> Environment:
    env = Environment(
        loader=FileSystemLoader(Path(__file__).parent / "templates"),
        autoescape=select_autoescape(),
    )
    env.filters["asset_amount"] = _format_asset_amount
    env.filters["ratio"] = _format_ratio
    env.filters["usd"] = _format_usd
    return env


def _top_customers(ledger_path: Path, prices_path: Path) -> list[dict[str, Any]]:
    if not ledger_path.exists():
        return []
    ledger = load_liabilities(ledger_path)
    prices = load_prices(prices_path)
    exposures: list[dict[str, Any]] = []
    for row in ledger.itertuples(index=False):
        balance = Decimal(str(row.balance))
        usd_value = balance * prices[str(row.asset)]
        exposures.append(
            {
                "customer_id": str(row.customer_id),
                "asset": str(row.asset),
                "balance": balance,
                "usd_value": usd_value,
            }
        )
    return sorted(exposures, key=lambda item: item["usd_value"], reverse=True)[:10]


def _write_summary(sheet: Any, rows: list[ReconciliationRow], as_of: date) -> None:
    headers = [
        "Asset",
        "Customer Liabilities",
        "Customer Count",
        "On-Chain Reserves",
        "Delta",
        "Coverage Ratio",
        "USD Price",
        "USD Value at Risk",
        "Status",
    ]
    sheet.append([f"Customer Liability Coverage Reconciliation - {as_of.isoformat()}"])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A1"].font = Font(bold=True, size=14, color="0B1F3A")
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row.asset,
                _decimal_to_number(row.customer_liabilities),
                row.customer_count,
                _decimal_to_number(row.on_chain_reserves),
                _decimal_to_number(row.delta),
                None if row.coverage_ratio is None else _decimal_to_number(row.coverage_ratio),
                _decimal_to_number(row.usd_price),
                _decimal_to_number(row.usd_value_at_risk),
                row.status,
            ]
        )
    _style_header(sheet, 2)
    for row_idx in range(3, 3 + len(rows)):
        sheet.cell(row=row_idx, column=6).number_format = "0.000%"
        sheet.cell(row=row_idx, column=7).number_format = "$#,##0.00"
        sheet.cell(row=row_idx, column=8).number_format = "$#,##0.00"
        status_cell = sheet.cell(row=row_idx, column=9)
        status_cell.alignment = Alignment(horizontal="center")
        status_cell.font = Font(bold=True)
        status_cell.fill = _status_fill(str(status_cell.value))
    sheet.conditional_formatting.add(
        f"F3:F{2 + len(rows)}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=_fill("E7F5EE")),
    )
    sheet.conditional_formatting.add(
        f"F3:F{2 + len(rows)}",
        CellIsRule(operator="between", formula=["0.99", "0.999999"], fill=_fill("FFF4D6")),
    )
    sheet.conditional_formatting.add(
        f"F3:F{2 + len(rows)}",
        CellIsRule(operator="lessThan", formula=["0.99"], fill=_fill("FDE8E7")),
    )


def _write_liabilities(sheet: Any, ledger_path: Path) -> None:
    sheet.append(["customer_id", "asset", "balance", "as_of"])
    _style_header(sheet, 1)
    if not ledger_path.exists():
        return
    ledger = load_liabilities(ledger_path)
    for row in ledger.itertuples(index=False):
        sheet.append([row.customer_id, row.asset, str(row.balance), row.as_of])


def _write_reserves(sheet: Any, rows: list[ReconciliationRow]) -> None:
    sheet.append(["asset", "wallet_address", "balance"])
    _style_header(sheet, 1)
    for row in rows:
        for detail in row.reserve_details:
            sheet.append([detail.asset, detail.address, _decimal_to_number(detail.balance)])


def _write_methodology(sheet: Any) -> None:
    lines = [
        ("Purpose", "Portfolio demonstration of daily customer-liability reserve coverage."),
        ("Liabilities", "Synthetic customer balances aggregated by asset using DuckDB SQL."),
        ("Reserves", "Public block-explorer wallet balances from mempool.space and Etherscan."),
        ("Coverage Ratio", "on_chain_reserves / customer_liabilities."),
        ("USD Value at Risk", "max(0, customer_liabilities - on_chain_reserves) * usd_price."),
        ("Thresholds", "OK >= 100%; WATCH >= 99% and < 100%; BREACH < 99%."),
        ("Disclaimer", "Synthetic ledger; public demo wallets; not an audit of any real exchange."),
    ]
    sheet.append(["Topic", "Methodology"])
    _style_header(sheet, 1)
    for line in lines:
        sheet.append(list(line))
        sheet.cell(row=sheet.max_row, column=2).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )


def _style_header(sheet: Any, row_idx: int) -> None:
    for cell in sheet[row_idx]:
        cell.fill = _fill("F4F7FB")
        cell.font = Font(bold=True, color="0B1F3A")
        cell.alignment = Alignment(horizontal="left")


def _autosize_columns(sheet: Any) -> None:
    for column_cells in sheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = min(max([len(value) for value in values] + [10]) + 2, 60)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _status_fill(status: str) -> PatternFill:
    return {
        "OK": _fill("E7F5EE"),
        "WATCH": _fill("FFF4D6"),
        "BREACH": _fill("FDE8E7"),
    }.get(status, _fill("FFFFFF"))


def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def _decimal_to_number(value: Decimal) -> int | float:
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def _format_asset_amount(value: Decimal, asset: str) -> str:
    places = 2 if asset in {"USDC", "USDT"} else 8
    return f"{value:,.{places}f}"


def _format_ratio(value: Decimal | None) -> str:
    if value is None:
        return "n/a"
    return f"{value * Decimal('100'):.3f}%"


def _format_usd(value: Decimal) -> str:
    return f"${value:,.2f}"
