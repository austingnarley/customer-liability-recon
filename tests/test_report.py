from __future__ import annotations

from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from src.reconcile import ReconciliationRow, ReserveDetail
from src.report import render_html, render_xlsx


def test_render_html_contains_key_sections(tmp_path, fixtures_dir) -> None:
    out_path = tmp_path / "report.html"

    render_html(
        [_sample_row()],
        date(2026, 5, 5),
        out_path,
        ledger_path=fixtures_dir / "sample_ledger.csv",
        prices_path=_prices_path(tmp_path),
    )

    html = out_path.read_text(encoding="utf-8")
    assert "Executive Summary" in html
    assert "Per-Asset Reserve Detail" in html
    assert "Top Customers by Exposure" in html
    assert "portfolio demonstration" in html


def test_render_xlsx_creates_expected_sheets_and_values(tmp_path, fixtures_dir) -> None:
    out_path = tmp_path / "report.xlsx"

    render_xlsx(
        [_sample_row()],
        date(2026, 5, 5),
        out_path,
        ledger_path=fixtures_dir / "sample_ledger.csv",
        prices_path=_prices_path(tmp_path),
    )

    workbook = load_workbook(out_path)
    assert workbook.sheetnames == ["Summary", "Liabilities", "Reserves", "Methodology"]
    assert workbook["Summary"].freeze_panes == "A3"
    assert workbook["Liabilities"].freeze_panes == "A2"
    assert workbook["Summary"]["A3"].value == "BTC"
    assert workbook["Summary"]["I3"].value == "OK"
    assert workbook["Reserves"]["A2"].value == "BTC"


def _sample_row() -> ReconciliationRow:
    return ReconciliationRow(
        asset="BTC",
        customer_liabilities=Decimal("100"),
        customer_count=10,
        on_chain_reserves=Decimal("101"),
        delta=Decimal("1"),
        coverage_ratio=Decimal("1.01"),
        usd_price=Decimal("67500"),
        usd_value_at_risk=Decimal("0"),
        status="OK",
        reserve_details=[
            ReserveDetail(
                asset="BTC",
                address="1A1zP1eP5QGefi2DMPTfTL5SLmv7DivfNa",
                balance=Decimal("101"),
            )
        ],
    )


def _prices_path(tmp_path):
    path = tmp_path / "prices.csv"
    path.write_text(
        """
asset,usd_price,as_of
BTC,67500.00,2026-05-05
ETH,3100.00,2026-05-05
USDC,1.00,2026-05-05
USDT,1.00,2026-05-05
""".strip(),
        encoding="utf-8",
    )
    return path
